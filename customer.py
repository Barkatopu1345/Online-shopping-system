import openpyxl
from online_shopping import *

def mailValidation(email):
    if '@' in email and '.' in email:
        return True
    else: return False


def createAccount():
    path = "shopingFile.xlsx"
    wb_obj = openpyxl.load_workbook(path)
    sheet = wb_obj['userAccount']
    row = sheet.row_max

    lst = []
    for i in range(2,row+1):
        lst.append(sheet.cell(i,2).value)

    name = input("Enter name: ")
    userID = input("UserId(name+digits. Ex: barkat1345): ")
    if userID in lst:
        print("User id can not be accepted! please try with another user id!")
        createAccount()
    password = input("Choose a password: ")
    address = input("Enter your address: ")
    phone = input("Enter your phone number: ")

    if(mailValidation(email) == False):
        email = input("Enter valid email: ")
  

    sheet['A1'] = "Name"
    sheet['B1'] = "User ID"
    sheet['C1'] = "Password"
    sheet['D1'] = "Address"
    sheet['E1'] = "Phone"

    sheet.cell(row+1,1,value = name)
    sheet.cell(row+1,2,value = userID)
    sheet.cell(row+1,3,value = password)

    wb_obj.save("shopingFile.xlsx")
    wb_obj.close()


def buyProducts():
    path = "shopingFile.xlsx"
    wb_obj = openpyxl.load_workbook(path)
    sheet = wb_obj['items']

    row = sheet.max_row

    printItems()
    productID = int(input("Enter product ID: "))
    quantity = int(input("Enter quantity: "))
    productName = "0"
    price = 0
    total = 0
    lst = []
    for i in range(1,row+1):
        lst.append(sheet.cell(i,1).value)
    if productID in lst:
        index = (lst.index(productID))+1
        if(sheet.cell(index,4).value<quantity):
            print("Does not stock this much products!")
            buyProducts()
        else:
            productName = sheet.cell(index,2).value
            price = sheet.cell(index,3).value
            total =sheet.cell(index,3).value * quantity 
    else:
        print("Invalid product id. Please reenter product id again!")
        buyProducts()

    wb_obj.save("shopingFile.xlsx")

    decission = input("Complete your parchase? y/n: ")
    if decission == 'y' or decission == 'Y':
        confirmBuy(productID,productName,price,quantity,total)
    

    wb_obj.close()



def confirmBuy(productID,productName,price,quantity,total):
    
    path = "shopingFile.xlsx"
    wb_obj = openpyxl.load_workbook(path)
    sheet = wb_obj['soldProduct']

    row = sheet.max_row
    column = sheet.max_column

    name = input("Enter your name: ")
    address = input("Enter your address: ")
    phone = input("Enter your phone number: ")
    method = input("Enter payment method: cash on delivary(COD)/bkash/nogod: ")    

    sheet['A1'] = "Product ID"
    sheet['B1'] = "Product Name"
    sheet['C1'] = "Quantity"
    sheet['D1'] = "Unit price"
    sheet['E1'] = "Total"
    sheet['F1'] = "Name"
    sheet['G1'] = "address"
    sheet['H1'] = "phone"
    sheet['I1'] = "Payment Method"

    print("Product ID "+" Product Name "+" Quantity "+" Unit price "+ " Total "+" Name "+" address "+" phone "+" Payment Method ")
    print(str(productID)+"        "+productName+"       "+str(quantity)+"  "+str(price)+"  "+str(total)+"  "+name+"  "+address+"  "+str(phone)+" "+method)
   
    decission = input("Confirm your order? y/n: ")
    if decission == 'y' or decission == 'Y':
        
        sheet.cell(row+1,1,value = productID)
        sheet.cell(row+1,2,value = productName)
        sheet.cell(row+1,3,value = quantity)
        sheet.cell(row+1,4,value = price)
        sheet.cell(row+1,5,value = total)
        sheet.cell(row+1,6,value = name)
        sheet.cell(row+1,7,value = address)
        sheet.cell(row+1,8,value = str(phone))
        sheet.cell(row+1,9,value = method)

    wb_obj.save("shopingFile.xlsx")
    wb_obj.close()
    print("Congratulations! ")




def logIn():
    wb_obj = openpyxl.load_workbook("shopingFile.xlsx")
    sheet = wb_obj['soldProduct']


    decission = input("Do your have any account? y/n: ")
    userId = "0"
    password = "0"
    if decission == 'y' or decission == 'Y':
        userId = input("Enter your user id: ")
        password = input("Enter your password: ")
    else:createAccount()



