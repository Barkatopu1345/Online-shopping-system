import openpyxl #for writting and reading excel file
from tabulate import tabulate # for generate table in printItems() and showItems()


productID = 1000  #for unique product id in the list



def giveID(): #creating unique id for the product
    path = "shopingFile.xlsx" #path of the excel file
    wb_obj = openpyxl.load_workbook(path) #creating openpyxl object by giving file path 
    sheet = wb_obj['items'] #sheet name from the shopingFile.xlsx
    lst = []
    
    __row__ = sheet.max_row #getting number of rows in the given list by using sheet.max_row 
    
    for i in range(2,__row__):
        x = sheet.cell(i,1).value #getting first value from every single rows. i represent the row no.and 1 reperent the first cell
        lst.append(x) #appending those value in the list
    return max(lst) # returning the largest value of the list



def createAccount(): #creating new account and storing in the userAccount sheet in the shopingFile.xlsx file
    path = "shopingFile.xlsx"
    wb_obj = openpyxl.load_workbook(path)
    sheet = wb_obj['userAccount']
    
    row = sheet.max_row #getting number of rows in the given list by using sheet.max_row

    lst = []
    for i in range(2,row+1):
        lst.append(sheet.cell(i,2).value) #getting 2nd value from every single rows. i represent the row no.and 2 reperent the 2nd cell

    name = input("Enter name: ")
    userID = input("UserId(name+digits. Ex: barkat1345): ")
    password = input("Choose a password: ")
    address = input("Enter your address: ")
    phone = input("Enter your phone number: ")

    if userID in lst: #checking user given userID is exist or not in the userAccount sheet to create unique userID
        print("User id can not be accepted! please try with another user id!")
        createAccount() # userID exist in the list. That's why calling this function recursively to do this process again.

    else: #if given userID not found in the sheet.

        lst2 = []
        sheet['A1'] = "Name"
        sheet['B1'] = "User ID"
        sheet['C1'] = "Password"
        sheet['D1'] = "Address"
        sheet['E1'] = "Phone"

        sheet.cell(row+1,1,value = name) #putting the user give value in the cell of the userAccount sheet of the excel file
        sheet.cell(row+1,2,value = userID) #cell() is the method of the openpyxl. "row" represent the number of rows occupied already.
        sheet.cell(row+1,3,value = password) 
        sheet.cell(row+1,4,value = address)
        sheet.cell(row+1,5,value = phone)

        lst2.append(name) #appending users information to use later.
        lst2.append(userID)
        lst2.append(password)
        lst2.append(address)
        lst2.append(phone)
        print("Congratulations! you have created an account. ")
        
        wb_obj.save("shopingFile.xlsx") #saving sheet after manipulation. otherwise new data will not save in the excel file
        
        return lst2 #returning user informations
    
    wb_obj.close() #closing the object
     

def logIn(userId,password): #login function for all kind of users
    wb_obj = openpyxl.load_workbook("shopingFile.xlsx")
    sheet = wb_obj['userAccount'] #sheet where users informations stored

    row = sheet.max_row # getting maximum number of rows occupied by the values
    column = sheet.max_column # getting maximum number of columns occupied by the values



    temp = []


    lst = []

    for i in range(2,row+2):
        lst.append(sheet.cell(i,2).value) #getting and appending only all userid and password in the list from the sheet.
        lst.append(sheet.cell(i,3).value)
    if userId not in lst or password not in lst: # checking users validation by checking userID and password
        print("Please enter valid user id or password! ")
        return False    
    else: #if found the id and password
        
        return True

    

def buyProducts(): # customer buy product
    path = "shopingFile.xlsx"
    wb_obj = openpyxl.load_workbook(path)
    sheet = wb_obj['items']

    row = sheet.max_row

    showItems() # showing all the items from the stock

    productID = int(input("Enter product ID: "))
    quantity = int(input("Enter quantity: "))


    productName = "0"
    price = 0
    total = 0
    lst = []

    for i in range(1,row+1):
        lst.append(sheet.cell(i,1).value) # getting and appending only all product id from the stock

    if productID in lst: # checking given product id valid or not
        index = (lst.index(productID))+1 # finding in which index in the list the desired product id found.
                                        # Because of excel file start from 1, updating index plus 1

        if(sheet.cell(index,4).value<quantity): # checking does the given quantity by the customer have stock or not
            print("Does not stock this much products! try again later")
            return 
        else: # if have enough product quantity in the stock 
            
            productName = sheet.cell(index,2).value
            price = sheet.cell(index,3).value
            total =sheet.cell(index,3).value * quantity
            stockQuantity = sheet.cell(index,4).value # getting stock quantity from the excel file

            sheet.cell(index,4,value = stockQuantity - quantity) # updating the quantity value in the excel file
            wb_obj.save("shopingFile.xlsx")

    else:
        print("Invalid product id. Please try again later!")
        return


    decission = input("Complete your parchase? y/n: ")
    if decission == 'y' or decission == 'Y':
        wb_obj.close()

        confirmBuy(productID,productName,price,quantity,total) #calling the confirmBuy() for confirmation and storing the purchase data in excel file




def confirmBuy(productID,productName,price,quantity,total): # purchase confirmation from this method
    
    path = "shopingFile.xlsx"
    wb_obj = openpyxl.load_workbook(path)
    sheet = wb_obj['soldProduct']

    row = sheet.max_row
    column = sheet.max_column

    print("For confirmation your purchase, login again")
    userId = input("Enter your user id: ")
    password = input("Enter your password: ")

    if logIn(userId,password) == False: #validate user's information 
        print("Please enter the valid id and password or you may do not have any account! ")
        return
    else:
        name = input("Enter your name: ")
        address = input("Enter your address: ")
        phone = input("Enter your phone number: ")
        method = input("Enter payment method: cash on delivary(COD)/bkash/nogod: ")  # asking payment method  

        sheet['A1'] = "Product ID" # naming the columns in the excel file
        sheet['B1'] = "Product Name"
        sheet['C1'] = "Quantity"
        sheet['D1'] = "Unit price"
        sheet['E1'] = "Total"
        sheet['F1'] = "Name"
        sheet['G1'] = "address"
        sheet['H1'] = "phone"
        sheet['I1'] = "Payment Method"
        sheet['J1'] = "Time"

        lst = [["Product ID",productID],["Product Name",productName],["Quantity",quantity],
            ["Unit price",price],[" Total",total],["Name",name],["address",address],
            ["phone",phone],[" Payment Method",method]
        ]
        


        print(tabulate(lst, tablefmt="grid")) # showing the product and user information before confirm the purchase by generating table by tabulate module

        decission = input("Confirm your order? y/n: ")
        if decission == 'y' or decission == 'Y':
            
            sheet.cell(row+1,1,value = productID) # writting the purchase information in the excel file
            sheet.cell(row+1,2,value = productName)
            sheet.cell(row+1,3,value = quantity)
            sheet.cell(row+1,4,value = price)
            sheet.cell(row+1,5,value = total)
            sheet.cell(row+1,6,value = name)
            sheet.cell(row+1,7,value = address)
            sheet.cell(row+1,8,value = str(phone))
            sheet.cell(row+1,9,value = method)



        wb_obj.save("shopingFile.xlsx") # saving the update
        wb_obj.close()
        print("Thank you for staying with us! ")



def writeItems(): # writting new product in the product/item list
    path = "shopingFile.xlsx"
    wb_obj = openpyxl.load_workbook(path)
    sheet = wb_obj['items']
    
    printItems()
    
    sheet['A1'] = "Product ID" # naming the column in excel sheet
    sheet['B1'] = "Product Name"
    sheet['C1'] = "Unit price"
    sheet['D1'] = "Quantity"

    __row__ = sheet.max_row
    column = sheet.max_column
    
    if __row__<2: # checking number of rows occupied by the products greater less then 2 or not. if less then 2, then
                 # no product information written in the item table. 
        id = productID # if no product found in the item table, then this is the first product information. And the the 
                       # product id will be 1000 by default.
    else: id =  sheet.cell(__row__,1).value # if found, then the last product id from the bottom of the item table stored in "id"
    
    i = __row__+1

    loopChecker = True

    while(loopChecker == True):
        
        name = input("Product Name: ")
        
        name = name.capitalize() # capitalizing the first alphabet of the product 


        if (checkExistingProduct(name) == True): # if given product id already stored in the item table,then the program stopped here
            x = input("Product already exist in the list!")
            loopChecker = False
            break

            
                
        elif checkExistingProduct(name) == False: # if not found similar product name in the item table       
            price = float(input("Enter product unit price(BDT): "))
            quantity = float(input("Quantity: "))
            id += 1 #creating new product id for unique product id
            
            
            
            if giveID()>= id: #rechecking. giveID() gives the largest product id from the excel file
                id = giveID() + 1
            
            sheet.cell(i,1,value = id) # writting the product information in the excel sheet

            sheet.cell(i,2,value = name)
            
            
            sheet.cell(i,3,value = price)
            
            sheet.cell(i,4,value = quantity)
            
            check = input("Want to add more items? y/n: ")
            print(end = '\n')
            if check == 'n' or check == 'N':
                loopChecker = False # whenever loopChecker is False, the while loop break

    
    wb_obj.save('shopingFile.xlsx') 
    printItems()
    wb_obj.close()



def checkExistingProduct(name): # checking the given product name by admin exist or not in the item table 
   
    path = "shopingFile.xlsx"
    wb_obj = openpyxl.load_workbook(path)
    sheet = wb_obj['items']

    __row__ = sheet.max_row

    lst = []
    
    for i in range(1,__row__+1):
        lst.append(sheet.cell(i,2).value) # getting all the product name from the excel sheet, and appending in the list
    
    if name in lst: # checking the given product name exist or not
        return True
    else: return False
    wb_obj.close()


def printItems(): # printing the item table from the excel sheet

    path = "shopingFile.xlsx"
    wb_obj = openpyxl.load_workbook(path)
    sheet = wb_obj['items']

    __row__ = sheet.max_row
    column = sheet.max_column

    head = [str(sheet.cell(1,1).value),str(sheet.cell(1,2).value),str(sheet.cell(1,3).value),str(sheet.cell(1,4).value)]
    #all the column headers store in the head list
    lst = []
    
    for i in range(2,__row__+1):
        temp = []

        temp.append(sheet.cell(i,1).value) # appending the cells value(product id,name,price etc)
        temp.append(sheet.cell(i,2).value)
        temp.append(sheet.cell(i,3).value)
        temp.append(sheet.cell(i,4).value)

        lst.append(temp) # appending "temp" list in the "lst" list

    print(tabulate(lst, headers=head, tablefmt="grid")) # generating table by tabulate module








def showItems():
    path = "shopingFile.xlsx"
    wb_obj = openpyxl.load_workbook(path)
    sheet = wb_obj['items']

    __row__ = sheet.max_row
    column = sheet.max_column

    head = [str(sheet.cell(1,1).value),str(sheet.cell(1,2).value),str(sheet.cell(1,3).value),"Stock"]
    #all the column headers store in the head list
    lst = []

    for i in range(2,__row__+1):
        temp = []

        temp.append(sheet.cell(i,1).value) # appending the cells value(product id,name,price etc)
        temp.append(sheet.cell(i,2).value)
        temp.append(sheet.cell(i,3).value)

        x = sheet.cell(i,4).value

        if x>0: # checking product quantity in the excel sheet is greater than 0 or not
            value = "in stock"
        else: value = "out of stock"
        
        temp.append(value)

        lst.append(temp) # appending "temp" list in the "lst" list

    print(tabulate(lst, headers=head, tablefmt="grid")) # generating table by tabulate module



def updateItems(): #updating existing product information of the item table 
    path = "shopingFile.xlsx"
    wb_obj = openpyxl.load_workbook(path)
    sheet = wb_obj['items']
   
    __row__ = sheet.max_row
    column = sheet.max_column
    
    lst = []
    for i in range(2,__row__+1): # getting and appending only product id in the list from the sheet
        x = sheet.cell(i,1).value
        lst.append(x)
    
    printItems()

    check = True
    while(check == True):
        id = int(input("Enter your product ID:" ))
        id = (lst.index(id))+2 # finding the proudct id from the list
        
        name = input("Product Name: ")
        
        sheet.cell(id,2,value = name) # updating the product name
        
        price = float(input("Enter product unit price(BDT): "))
        
        sheet.cell(id,3,value = price)# updating the product price
        quantity = float(input("Quantity: "))
        
        sheet.cell(id,4,value = quantity)# updating the product quantity
        
        wb_obj.save('shopingFile.xlsx') #updating the item table by saving

        check = input("Want to update more? y/n: ")
        if check == 'y' or check == 'Y':
            updateItems() # if user want to update more, calling this function again recursively
        else: check == False
    
    
    printItems() 
    wb_obj.close()


def deleteItems(): # delete an item from the sheet by admin
    path = "shopingFile.xlsx"
    wb_obj = openpyxl.load_workbook(path)
    sheet = wb_obj['items']

    __row__ = sheet.max_row

    printItems() #showing all the existing items in the list

    id = int(input("Enter the product ID: "))
    
    index = 0

    for i in range(1,__row__+1): #finding the index of the item id in the sheet
        if sheet.cell(i,1).value == id:  # if cells value and given id matched
            index = i # then index = i

    sheet.delete_rows(index,1) # Deleting the entire row, where the id was found.

    wb_obj.save('shopingFile.xlsx') #saving the work
    
    printItems()

    check = input("Want to delete more? y/n: ")
    if check == 'y' or check == 'Y':
        deleteItems() # if user want to delete more, calling this function again recursively
    else: return
    wb_obj.close()


decission = 00

def Menu(): # this function for calling other user dependent other functions.
    account = input("Do you have any account? y/n: ")
    if account == 'y' or account == 'Y':
        global decission # global variable
        existAccount = input("Do you want to login? y/n: ")
        
        if existAccount == 'y' or existAccount == 'Y':
            userId = input("Enter your user id: ")
            password = input("Enter your password: ")
            profile = logIn(userId,password) # getting user's information from logIn()
            
            if (profile == True and userId == 'barkat1345' and password == '1234'): # checking user admin or not
                print("Add new item in the list -> 1")  # admin userID = barkat1345, password = 1234
                print("Update an item in the list -> 2")
                print("Delete an item from the list -> 3")
                decission = int(input("Chose your option: "))
                return decission

            elif(profile == True and userId != 'barkat1345' and password != '1234'): # if user is not an admin
                print("Buy a product -> 4")
                decission = int(input("Chose your option: "))
                return decission
            elif(profile == False):
                print("You may do not have any account! or you enter the wrong id and password. Please try again")
        else: # if user don't want to log in
            
            showItems()

            print("Thank you for staying with us! See you later")
            
            return
    else: # if user does not have any account
        print("Creating new account!") 
        createAccount()  # creating new account
        return      





while(True):
    x = Menu()


    if x == 1:
        writeItems()
    elif x == 2:
        updateItems()
    elif x == 3:
        deleteItems()
    elif x == 4:
        buyProducts()
    decission = input("Do you want to logout? y/n: ")
    if decission == 'y' or decission == 'Y':
        break
# createAccount()
# print(logIn())
# printItems()
